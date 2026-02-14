from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from datetime import datetime
import quantstats as qs
import streamlit as st
import pandas as pd
import tempfile
import os
import io

def report(overview_df, risk_analysis_df, fundamentals_df, div_df, pf_returns, metrics):
        st.markdown("<h2 style='text-align:center; color:#7161ef;'>Export Portfolio Reports</h2>", unsafe_allow_html=True)
        st.markdown("<p style='text-align:center; color:#0096c7;'>Download detailed portfolio analytics in Excel or QuantStats report formats.</p>", unsafe_allow_html=True)
        
        all_dfs = [overview_df, risk_analysis_df, fundamentals_df, div_df]

        if all(df is None or df.empty for df in all_dfs):
            st.warning("No portfolio data available to generate report.")
            return
        
        with st.expander("Select Report Components", expanded=False):
            include_overview = st.checkbox("Include Overview", value=True)
            include_risk = st.checkbox("Include Risk Analysis", value=True)
            include_fundamentals = st.checkbox("Include Fundamentals", value=True)
            include_dividends = st.checkbox("Include Dividends", value=True)

        try:
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                include_sheets = []

                report_title = "PORTFOLIO ANALYTICS REPORT"
                generated_on = datetime.now().strftime("%d %B %Y, %H:%M")
                total_holdings = len(overview_df)
                
                gain = metrics.get("pf_gain_loss")
                cumulative_return = metrics.get("cumulative_return")
                cagr = metrics.get("cagr")
                ann_vol = metrics.get("volatility")
                sharpe = metrics.get("sharpe")
                sortino = metrics.get("sortino")
                max_dd = metrics.get("max_dd")
                
                top_summary = [ ["Generated On", generated_on],
                                ["Total Holdings", total_holdings]]
                top_summary = pd.DataFrame(top_summary, columns=["Metric", "Value"])
                top_summary.to_excel(writer, index=False, sheet_name="Report Info", startrow=2)

                summary = [ ["Portfolio Gain / Loss", f"₹{gain:.2f}"],
                            ["Cumulative Return", f"{cumulative_return:.2%}"],
                            ["Portfolio CAGR", f"{cagr:.2%}"],
                            ["Annualized Volatility", f"{ann_vol:.2%}"],
                            ["Sharpe Ratio", f"{sharpe:.2f}"],
                            ["Sortino Ratio", f"{sortino:.2f}"],
                            ["Max Drawdown", f"{abs(max_dd):.2%}"],]

                summary = pd.DataFrame(summary, columns=["Metric", "Value"])
                summary.to_excel(writer, index=False, sheet_name="Report Info", startrow=6)
                include_sheets.append("Report Info")

                if include_overview and overview_df is not None and not overview_df.empty:
                    overview_df.to_excel(writer, index=False, sheet_name="Performance Overview")
                    include_sheets.append("Performance Overview")

                if include_risk and risk_analysis_df is not None and not risk_analysis_df.empty:
                    risk_analysis_df.to_excel(writer, index=False, sheet_name="Risk Analysis")
                    include_sheets.append("Risk Analysis")

                if include_fundamentals and fundamentals_df is not None and not fundamentals_df.empty:
                    fundamentals_df.to_excel(writer, index=False, sheet_name="Fundamentals")
                    include_sheets.append("Fundamentals")

                if include_dividends and div_df is not None and not div_df.empty:
                    div_df.to_excel(writer, index=False, sheet_name="Dividends")
                    include_sheets.append("Dividends")

                for sheet in writer.sheets.values():
                    sheet.freeze_panes = "A2"

                    for column_cells in sheet.columns:
                        max_length = 0
                        col_letter = get_column_letter(column_cells[0].column)

                        for cell in column_cells:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))

                        sheet.column_dimensions[col_letter].width = max_length + 3

                sheet_r = writer.sheets["Report Info"]

                sheet_r.merge_cells("A1:B1")
                sheet_r["A1"] = report_title

                sheet_r["A1"].font = Font(size=18, bold=True)
                sheet_r["A1"].alignment = Alignment(horizontal="center", vertical="center")
                sheet_r.row_dimensions[1].height = 30

                for row in sheet_r.iter_rows(min_row=3, max_row=5, min_col=1, max_col=1):
                    for cell in row:
                        cell.font = Font(bold=True)
                
                for cell in sheet_r["A"]:
                    cell.font = cell.font.copy(bold=True)
                
            buffer.seek(0)
            st.download_button("📘 Download Excel Report", data=buffer.getvalue(), file_name=f"Portfolio_Report_{datetime.now().strftime("%d%m%Y")}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            st.warning(f"Could not export Excel: {e}")

        if not pf_returns.empty:
            try:
                tmpdir = tempfile.TemporaryDirectory()
                report_path = os.path.join(tmpdir.name, "portfolio_report.html")
                qs.reports.html(pf_returns, output=report_path, title="Portfolio Performance Report")
                with open(report_path, "r", encoding="utf-8") as f:
                    html_data = f.read()
                st.download_button("📊 Download Performance Report (QuantStats)", data=html_data, file_name=f"Portfolio_Report_{datetime.now().strftime('%Y%m%d')}.html", mime="text/html")
                tmpdir.cleanup()
            except Exception as e:
                st.warning(f"Could not generate QuantStats report: {e}")

